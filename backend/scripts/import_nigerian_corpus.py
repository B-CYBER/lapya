"""Generate `app/seeds/recipes.py` from the Nigerian corpus spreadsheets.

Run:  ./venv/bin/python -m scripts.import_nigerian_corpus

Reads:
  data/nigerian_corpus/food_database.xlsx  (Cooked Foods by Region — the backbone)
  data/nigerian_corpus/diet_guide.xlsx     (per-condition FOODS TO EAT / AVOID)
  backend/static/recipe-images/            (slugged photos, written by process_recipe_images.py)

Writes a deterministic, diffable Python module: app/seeds/recipes.py

Condition safety is derived two ways, both grounded in the diet guide:
  1. Direct match — dish name matches a guide "FOODS TO EAT" (safe) or
     "FOODS TO AVOID" entry (avoid, or caution when the entry is qualified
     with "large portions"/"excess"/etc.).
  2. Rule layer — generalises the guide's category statements to dish
     classes the guide names only generically (e.g. "Eba/Garri large
     portions = high refined carbs" → all swallows are diabetes-caution).
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl

from scripts.corpus_common import (
    FOOD_DB,
    DIET_GUIDE,
    IMAGE_OUT_DIR,
    normalize_food_key,
    normalize_tokens,
    parse_calories,
    slugify,
)

OUTPUT = Path(__file__).resolve().parents[1] / "app" / "seeds" / "recipes.py"

ZONE_TO_REGION = {
    "north west": "north-west",
    "north east": "north-east",
    "north central": "north-central",
    "south west": "south-west",
    "south east": "south-east",
    "south south": "south-south",
}

# Which condition slugs each diet-guide sheet informs.
SHEET_CONDITIONS = {
    "1. Diabetes + Hypertension": ["diabetes", "hypertension"],
    "2. Hypertension + CKD": ["hypertension", "kidney"],
    "3. Triple (DM+HTN+CKD)": ["diabetes", "hypertension", "kidney"],
    "4. High Uric Acid (Gout)": ["uric-acid"],
    "5. High Cholesterol": ["cholesterol", "heart"],
}

# A qualified "avoid" (e.g. "Eba/Garri (large portions)") is really a caution.
SOFT_QUALIFIERS = ("large portion", "excess", "moderate", "small portion", "limit", "occasional")

ALL_CONDITIONS = ["diabetes", "hypertension", "kidney", "heart", "cholesterol", "uric-acid", "weight"]


# ---------------------------------------------------------------------------
# Diet-guide keyset
# ---------------------------------------------------------------------------
def build_guide_keysets() -> tuple[dict[str, list[str]], dict[str, list[tuple[str, bool]]]]:
    wb = openpyxl.load_workbook(DIET_GUIDE, data_only=True)
    eat: dict[str, list[str]] = {}
    avoid: dict[str, list[tuple[str, bool]]] = {}
    for sheet, conditions in SHEET_CONDITIONS.items():
        ws = wb[sheet]
        section: str | None = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row[1:5] if c]
            line = " ".join(cells)
            if "✅" in line:
                section = "eat"
                continue
            if "❌" in line or "🚫" in line:
                section = "avoid"
                continue
            if not cells or section is None:
                continue
            if line.lower().startswith(("category", "food to avoid")):
                continue
            # EAT rows are Category | Food | Why; AVOID rows are Food | Reason.
            food = cells[1] if (section == "eat" and len(cells) >= 3) else cells[0]
            key = normalize_food_key(food)
            if len(key) < 4:
                continue
            soft = any(q in food.lower() for q in SOFT_QUALIFIERS)
            for condition in conditions:
                if section == "eat":
                    eat.setdefault(condition, []).append(key)
                else:
                    avoid.setdefault(condition, []).append((key, soft))
    return eat, avoid


def _keymatch(dish_key: str, guide_key: str) -> bool:
    if dish_key == guide_key:
        return True
    short, long = sorted((dish_key, guide_key), key=len)
    return len(short) >= 4 and short in long and len(short) >= 0.6 * len(long)


# ---------------------------------------------------------------------------
# Rule layer — generalises guide category statements to dish classes
# ---------------------------------------------------------------------------
# Each rule: (name keywords, category keywords) -> {condition: rating}
# Grounded in the diet guide's own statements; see module docstring.
SWALLOW_WORDS = ("eba", "garri", "pounded yam", "iyan", "amala", "semo", "semovita",
                 "fufu", "akpu", "starch", "tuwo", "pupuru", "wheat swallow")
SWALLOW_SAFE_WORDS = ("oat", "plantain", "guinea corn", "sorghum", "wheat fufu")  # guide-recommended
OFFAL_WORDS = ("nkwobi", "isi ewu", "isiewu", "kpomo", "ponmo", "ugba", "ukpaka",
               "abacha", "offal", "cow leg", "cow foot")
CURED_MEAT_WORDS = ("kilishi", "suya", "tsire", "isire", "balangu", "dambu", "kundi")


def rule_layer(name: str, category: str, calories: int | None) -> dict[str, str]:
    text = f"{name} {category}".lower()
    out: dict[str, str] = {}

    is_swallow = any(w in text for w in SWALLOW_WORDS) and not any(
        w in text for w in SWALLOW_SAFE_WORDS
    )
    if is_swallow:
        # High refined carb / high glycaemic — guide cautions eba/pounded yam/white rice.
        out["diabetes"] = "caution"

    if any(w in text for w in OFFAL_WORDS):
        # Guide: organ meats — high saturated fat, cholesterol, purines.
        out["cholesterol"] = "avoid"
        out["heart"] = "avoid"
        out["uric-acid"] = "avoid"

    if any(w in text for w in CURED_MEAT_WORDS):
        # Cured/grilled spiced meat — high sodium + purine.
        out.setdefault("hypertension", "caution")
        out.setdefault("uric-acid", "caution")

    # Weight management is portion/calorie control (not a guide sheet).
    if calories is not None:
        if calories >= 700:
            out["weight"] = "avoid"
        elif calories >= 500:
            out["weight"] = "caution"

    return out


_SEVERITY = {"safe": 0, "caution": 1, "avoid": 2}


def derive_condition_safety(
    name: str,
    alt: str | None,
    category: str,
    calories: int | None,
    eat: dict[str, list[str]],
    avoid: dict[str, list[tuple[str, bool]]],
) -> dict[str, str]:
    dish_keys = [normalize_food_key(name)]
    if alt:
        for part in str(alt).split("/"):
            k = normalize_food_key(part)
            if len(k) >= 4:
                dish_keys.append(k)

    result: dict[str, str] = {}
    for condition in ALL_CONDITIONS:
        rating = "safe"
        for guide_key, soft in avoid.get(condition, []):
            if any(_keymatch(dk, guide_key) for dk in dish_keys):
                rating = "caution" if soft else "avoid"
                break
        result[condition] = rating

    # Merge rule layer, keeping the stricter rating.
    for condition, rating in rule_layer(name, category, calories).items():
        if _SEVERITY[rating] > _SEVERITY.get(result.get(condition, "safe"), 0):
            result[condition] = rating

    # Drop "safe" entries — they're the read-time default, keeps JSON compact.
    return {c: r for c, r in result.items() if r != "safe"}


# ---------------------------------------------------------------------------
# Meal-type classification
# ---------------------------------------------------------------------------
def classify_meal_type(name: str, category: str) -> str:
    cat = (category or "").lower()
    words = set(normalize_tokens(name).split())
    if "breakfast" in cat or words & {"pap", "ogi", "akamu", "kunu", "akara", "masa", "funkaso", "fura", "kunun"}:
        return "breakfast"
    if any(w in cat for w in ("snack", "street food", "bbq", "drink", "beverage", "delicacy", "appetizer", "salad")):
        return "snack"
    if "rice" in cat or "rice" in words:
        return "lunch"
    if "porridge" in cat or words & {"asaro", "adalu", "pottage", "ikokore"}:
        return "lunch"
    if any(w in cat for w in ("soup", "stew", "sauce")):
        return "dinner"
    if any(w in cat for w in ("swallow", "staple", "combo", "side")):
        return "dinner"
    return "dinner"


# ---------------------------------------------------------------------------
# Ingredient vocabulary (mined from the corpus' own sheets)
# ---------------------------------------------------------------------------
def build_ingredient_vocab() -> list[tuple[str, str, str]]:
    """Return (match_token, display_name, grocery_category) tuples."""
    wb = openpyxl.load_workbook(FOOD_DB, data_only=True)
    vocab: list[tuple[str, str, str]] = []

    veg = wb["Vegetables"]
    for row in veg.iter_rows(min_row=3, values_only=True):
        name = row[1]
        if not name:
            continue
        vocab.append((normalize_tokens(str(name).split("(")[0]), str(name).split("(")[0].strip(), "vegetables"))

    ING_CAT = {
        "oil": "essentials", "fermented condiment": "essentials", "spice": "essentials",
        "seasoning": "essentials", "pepper": "vegetables", "grain": "grains",
        "tuber": "grains", "legume": "proteins", "protein": "proteins",
    }
    ing = wb["Cooking Ingredients"]
    for row in ing.iter_rows(min_row=3, values_only=True):
        name, _local, cat = row[1], row[2], row[3]
        if not name:
            continue
        base = str(name).split("(")[0].strip()
        gcat = "essentials"
        if cat:
            for key, mapped in ING_CAT.items():
                if key in str(cat).lower():
                    gcat = mapped
                    break
        vocab.append((normalize_tokens(base), base, gcat))

    # De-dup by token, keep longest tokens first so multi-word matches win.
    seen: set[str] = set()
    uniq: list[tuple[str, str, str]] = []
    for token, display, gcat in sorted(vocab, key=lambda t: -len(t[0])):
        if len(token) < 3 or token in seen:
            continue
        seen.add(token)
        uniq.append((token, display, gcat))
    return uniq


# Per-meal-type staple padding so every recipe has >= 3 ingredients.
STAPLE_PADDING = {
    "dinner": [("Palm oil", "1 tbsp", "essentials"), ("Fresh pepper", "to taste", "vegetables"),
               ("Onion", "1 medium", "vegetables"), ("Smoked fish", "1 portion", "proteins")],
    "lunch": [("Onion", "1 medium", "vegetables"), ("Tomato", "2 medium", "vegetables"),
              ("Vegetable oil", "1 tbsp", "essentials")],
    "breakfast": [("Groundnut", "1 handful", "proteins"), ("Milk", "to taste", "essentials")],
    "snack": [("Groundnut oil", "for frying", "essentials"), ("Salt", "a pinch", "essentials")],
}


def mine_ingredients(name: str, description: str, meal_type: str,
                     vocab: list[tuple[str, str, str]]) -> list[tuple[str, str, str]]:
    haystack = normalize_tokens(f"{name} {description or ''}")
    found: list[tuple[str, str, str]] = []
    used: set[str] = set()
    for token, display, gcat in vocab:
        if token in used:
            continue
        if f" {token} " in f" {haystack} ":
            found.append((display, "1 portion", gcat))
            used.add(token)
        if len(found) >= 6:
            break
    # Pad to a minimum of 3 with meal-type staples.
    for item, amount, gcat in STAPLE_PADDING.get(meal_type, []):
        if len(found) >= 4:
            break
        if item.lower() not in {f[0].lower() for f in found}:
            found.append((item, amount, gcat))
    return found[:6]


# ---------------------------------------------------------------------------
# Macro estimation (corpus gives calories only; macros are display-only)
# ---------------------------------------------------------------------------
def estimate_macros(calories: int, meal_type: str, category: str) -> dict[str, int]:
    cat = (category or "").lower()
    if any(w in cat for w in ("swallow", "staple", "rice", "porridge", "bread")):
        split = (0.10, 0.72, 0.18)  # protein, carb, fat fractions of energy
        sodium, potassium = 60, 280
    elif any(w in cat for w in ("soup", "stew", "sauce")):
        split = (0.30, 0.20, 0.50)
        sodium, potassium = 480, 520
    elif any(w in cat for w in ("meat", "bbq", "street")):
        split = (0.35, 0.10, 0.55)
        sodium, potassium = 520, 360
    else:
        split = (0.18, 0.55, 0.27)
        sodium, potassium = 200, 320
    p, c, f = split
    return {
        "protein": round(calories * p / 4),
        "carbs": round(calories * c / 4),
        "fat": round(calories * f / 9),
        "sodium_mg": sodium,
        "potassium_mg": potassium,
    }


# ---------------------------------------------------------------------------
# Image matching
# ---------------------------------------------------------------------------
FALLBACK_IMAGE = {
    "breakfast": "pap-and-milk",
    "lunch": "boiled-rice-and-beans-with-sauce",
    "dinner": "vegetable-soup",
    "snack": "puff-puff",
}

# Generic words shared by many dishes — ignored when scoring token overlap so a
# match needs a *distinctive* word (egusi, onugbu, okra…), not just "soup".
IMAGE_STOPWORDS = {
    "soup", "rice", "stew", "sauce", "and", "with", "the", "fried", "boiled",
    "grilled", "roasted", "nigerian", "fresh", "style", "dish", "swallow", "pepper",
}

# Curated overrides for famous dishes whose photo filename doesn't token-match.
# Keyed by dish-slug prefix → image stem.
IMAGE_ALIASES = {
    "eba-garri": "garri-with-milk-and-groundnut",
    "amala": "amala-and-vegetable-soup",
    "iyan": "amala-and-vegetable-soup",
    "pounded-yam": "amala-and-vegetable-soup",
    "ogi-akamu-pap": "pap-and-milk",
    "kunu": "pap-and-milk",
    "boli": "bole-and-fish",
    "ofada-rice-ayamase": "ofada-sauce",
    "jollof-rice": "rice-and-vegetable-sauce",
    "edo-banga-rice": "rice-and-vegetable-sauce",
    "abakaliki-rice": "rice-and-vegetable-sauce",
    "nkwobi": "peppered-kpomo",
    "pepper-soup": "pepper-soup",
    "delta-peppersoup": "beef-pepper-soup",
}


def build_image_index() -> set[str]:
    if not IMAGE_OUT_DIR.exists():
        return set()
    return {p.stem for p in IMAGE_OUT_DIR.glob("*.jpg")}


def match_image(name: str, alt: str | None, meal_type: str, available: set[str]) -> str:
    slug = slugify(name)
    if slug in available:
        return f"/static/recipe-images/{slug}.jpg"
    for prefix, stem in IMAGE_ALIASES.items():
        if slug.startswith(prefix) and stem in available:
            return f"/static/recipe-images/{stem}.jpg"

    name_text = normalize_tokens(name)
    if alt:
        name_text += " " + normalize_tokens(str(alt))
    name_tokens = {t for t in name_text.split() if len(t) >= 4 and t not in IMAGE_STOPWORDS}

    best: str | None = None
    best_overlap = 0
    for stem in available:
        stem_tokens = {t for t in stem.replace("-", " ").split() if t not in IMAGE_STOPWORDS}
        overlap = len(name_tokens & stem_tokens)
        if overlap > best_overlap:
            best_overlap = overlap
            best = stem
    if best and best_overlap >= 1:
        return f"/static/recipe-images/{best}.jpg"
    return f"/static/recipe-images/{FALLBACK_IMAGE[meal_type]}.jpg"


# ---------------------------------------------------------------------------
# Steps
# ---------------------------------------------------------------------------
def build_steps(description: str | None) -> list[str]:
    if not description:
        return ["Prepare ingredients.", "Cook through.", "Serve warm."]
    parts = [p.strip() for p in str(description).replace(";", ".").split(".") if p.strip()]
    if len(parts) < 2:
        parts.append("Serve warm")
    steps = []
    for p in parts[:5]:
        p = p.rstrip(".")
        steps.append(p[:1].upper() + p[1:] + ".")
    return steps


PREP_COOK = {
    "breakfast": (15, 20), "lunch": (20, 35), "dinner": (25, 45), "snack": (15, 20),
}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def collect_recipes() -> list[dict]:
    eat, avoid = build_guide_keysets()
    vocab = build_ingredient_vocab()
    images = build_image_index()

    wb = openpyxl.load_workbook(FOOD_DB, data_only=True)
    ws = wb["Cooked Foods by Region"]

    recipes: list[dict] = []
    seen_slugs: set[str] = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, alt, zone, _state, ethnic, category, description, cal_raw = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]
        )
        if not name or str(name).startswith("Source"):
            continue
        name = str(name).strip()
        slug = slugify(name)
        if slug in seen_slugs:
            continue
        seen_slugs.add(slug)

        calories = parse_calories(cal_raw) or 300
        meal_type = classify_meal_type(name, str(category or ""))
        region = ZONE_TO_REGION.get(str(zone or "").strip().lower())
        macros = estimate_macros(calories, meal_type, str(category or ""))
        safety = derive_condition_safety(name, alt, str(category or ""), calories, eat, avoid)
        ingredients = mine_ingredients(name, str(description or ""), meal_type, vocab)
        steps = build_steps(description)
        prep, cook = PREP_COOK[meal_type]
        image = match_image(name, alt, meal_type, images)

        reason = (str(description or "").split(";")[0].split(".")[0].strip() or
                  f"Traditional {region or 'Nigerian'} dish.")[:200]

        recipes.append({
            "slug": slug,
            "name": name,
            "local_name": (str(alt).strip()[:120] if alt else None),
            "description": (str(description).strip()[:500] if description else None),
            "meal_type": meal_type,
            "image": image,
            "portion": "1 serving",
            "reason": reason,
            "prep": prep, "cook": cook, "servings": 1,
            "calories": calories,
            "protein": macros["protein"], "carbs": macros["carbs"], "fat": macros["fat"],
            "sodium_mg": macros["sodium_mg"], "potassium_mg": macros["potassium_mg"],
            "region": region,
            "ethnic_origin": (str(ethnic).strip() if ethnic else None),
            "condition_safety": safety,
            "ingredients": [(i, a, c, 0) for (i, a, c) in ingredients],
            "steps": steps,
        })
    return recipes


SEED_FUNCTION = '''

def seed_recipes(db: Session) -> None:
    seen_slugs = {spec["slug"] for spec in RECIPES}

    # Drop any previously-seeded recipes that are no longer in the corpus,
    # including rows that reference them (SQLite does not enforce FKs by
    # default, so without this meal plans keep orphaned recipe_ids and the
    # dashboard 500s). Affected users get a fresh plan on next dashboard load.
    stale_recipes = db.scalars(
        select(Recipe).where(Recipe.is_seeded.is_(True), Recipe.slug.notin_(seen_slugs))
    ).all()
    if stale_recipes:
        stale_ids = [r.id for r in stale_recipes]
        db.execute(delete(MealPlanItem).where(MealPlanItem.recipe_id.in_(stale_ids)))
        db.execute(delete(ScanResult).where(ScanResult.recipe_id.in_(stale_ids)))
        for stale in stale_recipes:
            db.delete(stale)

    for spec in RECIPES:
        existing = db.scalar(select(Recipe).where(Recipe.slug == spec["slug"]))
        recipe = existing or Recipe(slug=spec["slug"])
        recipe.name = spec["name"]
        recipe.local_name = spec.get("local_name")
        recipe.description = spec.get("description")
        recipe.meal_type = spec["meal_type"]
        recipe.image_url = spec["image"]
        recipe.prep_minutes = spec["prep"]
        recipe.cook_minutes = spec["cook"]
        recipe.servings = spec["servings"]
        recipe.calories = spec["calories"]
        recipe.protein_g = spec["protein"]
        recipe.carbs_g = spec["carbs"]
        recipe.fat_g = spec["fat"]
        recipe.sodium_mg = spec["sodium_mg"]
        recipe.potassium_mg = spec["potassium_mg"]
        recipe.portion = spec["portion"]
        recipe.reason = spec["reason"]
        recipe.condition_safety = spec["condition_safety"]
        recipe.cuisine_region = spec.get("region")
        recipe.is_seeded = True

        if existing is None:
            db.add(recipe)
            db.flush()

        recipe.ingredients.clear()
        for item, amount, category, naira_kobo in spec["ingredients"]:
            recipe.ingredients.append(
                RecipeIngredient(
                    item=item, amount=amount, category=category,
                    naira_kobo=naira_kobo, sort_order=0,
                )
            )
        recipe.steps.clear()
        for index, instruction in enumerate(spec["steps"], start=1):
            recipe.steps.append(RecipeStep(step_number=index, instruction=instruction))

    db.commit()
'''


def render_module(recipes: list[dict]) -> str:
    header = f'''"""Nigerian therapeutic recipe corpus — AUTO-GENERATED, do not edit by hand.

Source: data/nigerian_corpus/{{food_database.xlsx, diet_guide.xlsx}}
Regenerate: ./venv/bin/python -m scripts.import_nigerian_corpus
Generated: {_dt.date.today().isoformat()}  ({len(recipes)} recipes)

condition_safety maps a condition slug to "caution" or "avoid"; a missing
slug means "safe" (the read-time default). Ratings derive from the
Nigerian Multi-Morbidity Diet Guide plus a documented rule layer.
"""

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.meal_plan import MealPlanItem
from app.models.recipe import Recipe, RecipeIngredient, RecipeStep
from app.models.scan import ScanResult

RECIPES: list[dict] = '''
    import pprint
    body = pprint.pformat(recipes, indent=4, width=100, sort_dicts=False)
    return header + body + "\n" + SEED_FUNCTION


def main() -> None:
    recipes = collect_recipes()
    OUTPUT.write_text(render_module(recipes))

    by_type: dict[str, int] = {}
    safety_hits = 0
    fallback_imgs = 0
    for r in recipes:
        by_type[r["meal_type"]] = by_type.get(r["meal_type"], 0) + 1
        if r["condition_safety"]:
            safety_hits += 1
        if "/vegetable-soup." in r["image"] or "/pap-and-milk." in r["image"] or \
           "/boiled-rice-and-beans-with-sauce." in r["image"] or "/puff-puff." in r["image"]:
            fallback_imgs += 1
    print(f"Wrote {len(recipes)} recipes -> {OUTPUT.relative_to(Path.cwd())}")
    print(f"  meal types: {by_type}")
    print(f"  with non-default safety: {safety_hits}/{len(recipes)}")
    print(f"  using fallback image: {fallback_imgs}/{len(recipes)}")


if __name__ == "__main__":
    main()
